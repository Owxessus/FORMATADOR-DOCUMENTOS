# -*- coding: utf-8 -*-
"""Assistente conversacional: conversa, lê anexos, gera imagens e edita
planilhas — tudo pela mesma chave da OpenRouter.

Segurança: a IA nunca executa código. Para mexer numa planilha ela devolve
uma LISTA DE OPERAÇÕES (escrever célula, fórmula, formato, largura…) que
este módulo aplica com openpyxl. Assim o que pode acontecer com o arquivo
é sempre um conjunto conhecido de ações.
"""
from __future__ import annotations

import base64
import datetime
import json
import os
import re
import zipfile

import requests

import ai_client
import settings

ARQ_CONVERSA = "conversa.json"
MODELO_IMAGEM = "bytedance-seed/seedream-5-0-lite"

SYSTEM = """Você é a assistente do Formatador de Relatórios, usada por uma \
gerente de serviço de acolhimento da assistência social. Responda em \
português do Brasil, de forma direta e prática.

Você pode:
- conversar e ajudar a redigir, resumir e revisar textos;
- consultar a web quando a busca estiver ativada (nesse caso, cite as fontes
  e não invente: se não achar, diga que não achou);
- ler o conteúdo dos anexos que aparecem na conversa;
- gerar imagens;
- editar planilhas Excel anexadas.

Para GERAR UMA IMAGEM, responda apenas com:
{"acao": "gerar_imagem", "prompt": "descrição detalhada em inglês"}

Para CRIAR UMA APRESENTAÇÃO, responda apenas com:
{"acao": "criar_apresentacao", "titulo": "...", "subtitulo": "...",
 "slides": [{"titulo": "...", "topicos": ["...", "..."], "notas": "..."}]}
Regras: 4 a 6 tópicos por slide, frases curtas, sem inventar dados.

Para EDITAR A APRESENTAÇÃO anexada:
{"acao": "editar_apresentacao", "resumo": "...", "operacoes": [
  {"tipo": "substituir_texto", "de": "2025", "para": "2026"},
  {"tipo": "alterar_titulo", "slide": 2, "valor": "..."},
  {"tipo": "adicionar_slide", "titulo": "...", "topicos": ["..."]},
  {"tipo": "remover_slide", "slide": 5},
  {"tipo": "notas", "slide": 1, "valor": "..."}]}

Para CONVERTER OU MANIPULAR ARQUIVOS anexados:
{"acao": "arquivo", "resumo": "...", "tipo": "<operação>", "parametros": {...}}
Operações: juntar_pdf · dividir_pdf · extrair_paginas (paginas: "1-3,7") ·
girar_pdf (graus, paginas) · proteger_pdf (senha) · pdf_para_docx ·
pdf_para_imagens · imagens_para_pdf · docx_para_pdf · docx_para_pptx ·
pptx_para_pdf

Sobre PDF, seja honesta: dá para juntar, dividir, girar, extrair páginas,
proteger e transcrever sem converter; mas EDITAR O TEXTO exige converter
para .docx primeiro (o layout não é preservado) — proponha isso quando ela
pedir para "editar" ou "corrigir" um PDF.

Para EDITAR A PLANILHA anexada, responda apenas com:
{"acao": "editar_planilha", "resumo": "o que será feito",
 "operacoes": [
   {"tipo": "escrever", "planilha": "Plan1", "celula": "D1", "valor": "Total"},
   {"tipo": "formula", "planilha": "Plan1", "celula": "D2", "valor": "=B2*C2"},
   {"tipo": "preencher_abaixo", "planilha": "Plan1", "celula": "D2", "ate_linha": 50},
   {"tipo": "formato", "planilha": "Plan1", "intervalo": "A1:D1",
    "negrito": true, "fundo": "FFF3B0", "centralizar": true},
   {"tipo": "largura", "planilha": "Plan1", "coluna": "A", "valor": 32},
   {"tipo": "congelar", "planilha": "Plan1", "celula": "A2"},
   {"tipo": "numero", "planilha": "Plan1", "intervalo": "C2:C50",
    "formato": "#,##0.00"}
 ]}

Regras das planilhas: nunca invente dados; use as colunas que existem; se \
faltar informação, pergunte antes em texto normal. Em qualquer outro caso, \
responda em texto comum (sem JSON)."""


# ------------------------------------------------------------ conversa

def _caminho() -> str:
    return os.path.join(settings._config_dir(), ARQ_CONVERSA)


def carregar() -> list[dict]:
    try:
        with open(_caminho(), encoding="utf-8") as f:
            return json.load(f)
    except (OSError, json.JSONDecodeError):
        return []


def salvar(msgs: list[dict]) -> None:
    with open(_caminho(), "w", encoding="utf-8") as f:
        json.dump(msgs[-200:], f, ensure_ascii=False, indent=1)


def limpar() -> None:
    salvar([])


# -------------------------------------------------------------- anexos

IMAGENS = (".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp")


def _texto_docx(caminho: str) -> str:
    with zipfile.ZipFile(caminho) as z:
        xml = z.read("word/document.xml").decode("utf-8", "ignore")
    partes = re.findall(r"<w:t(?:\s[^>]*)?>(.*?)</w:t>", xml, flags=re.S)
    txt = " ".join(partes)
    for e, c in (("&amp;", "&"), ("&lt;", "<"), ("&gt;", ">"),
                 ("&quot;", '"'), ("&apos;", "'")):
        txt = txt.replace(e, c)
    return re.sub(r"\s+", " ", txt).strip()


def _texto_xlsx(caminho: str, max_linhas: int = 80) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(caminho, data_only=False)
    saida = []
    for ws in wb.worksheets:
        saida.append(f"[Planilha: {ws.title}  ({ws.max_row} linhas × "
                     f"{ws.max_column} colunas)]")
        for linha in ws.iter_rows(min_row=1, max_row=min(ws.max_row,
                                                         max_linhas)):
            celulas = [f"{c.coordinate}={c.value}" for c in linha
                       if c.value is not None]
            if celulas:
                saida.append("  " + "; ".join(celulas))
        if ws.max_row > max_linhas:
            saida.append(f"  … (mostrando as {max_linhas} primeiras linhas)")
    return "\n".join(saida)


def _texto_pdf(caminho: str) -> str:
    try:
        from pypdf import PdfReader
    except ImportError:
        return "(não foi possível ler o PDF neste computador)"
    r = PdfReader(caminho)
    return "\n".join((p.extract_text() or "") for p in r.pages[:30]).strip()


def preparar_anexo(caminho: str) -> list[dict]:
    """Prepara o anexo para a conversa.

    Devolve uma LISTA porque um PDF digitalizado (sem texto) é convertido
    em imagens de página, para o modelo de visão fazer o OCR.
    """
    ext = os.path.splitext(caminho)[1].lower()
    nome = os.path.basename(caminho)

    if ext == ".pdf":
        import arquivos
        if arquivos.pdf_tem_texto(caminho):
            return [{"tipo": "texto", "nome": nome, "caminho": caminho,
                     "conteudo": _texto_pdf(caminho)}]
        paginas = arquivos.pdf_para_imagens(caminho, limite=6)
        saida = [{"tipo": "texto", "nome": nome, "caminho": caminho,
                  "conteudo": "(PDF digitalizado, sem texto: as páginas "
                              "abaixo vão como imagem para leitura)"}]
        for img in paginas:
            with open(img, "rb") as f:
                b64 = base64.b64encode(f.read()).decode()
            saida.append({"tipo": "imagem", "nome": os.path.basename(img),
                          "caminho": img,
                          "conteudo": f"data:image/png;base64,{b64}"})
        return saida

    if ext == ".pptx":
        import apresentacao
        return [{"tipo": "texto", "nome": nome, "caminho": caminho,
                 "conteudo": apresentacao.ler(caminho)}]

    if ext in IMAGENS:
        with open(caminho, "rb") as f:
            b64 = base64.b64encode(f.read()).decode()
        mime = "image/png" if ext == ".png" else "image/jpeg"
        return [{"tipo": "imagem", "nome": nome, "caminho": caminho,
                 "conteudo": f"data:{mime};base64,{b64}"}]
    if ext == ".docx":
        return [{"tipo": "texto", "nome": nome, "caminho": caminho,
                 "conteudo": _texto_docx(caminho)}]
    if ext == ".xlsx":
        return [{"tipo": "texto", "nome": nome, "caminho": caminho,
                 "conteudo": _texto_xlsx(caminho), "planilha": caminho}]
    with open(caminho, encoding="utf-8", errors="ignore") as f:
        return [{"tipo": "texto", "nome": nome, "caminho": caminho,
                 "conteudo": f.read()[:20000]}]


# ------------------------------------------------------------- conversa IA

def _mensagem_usuario(texto: str, anexos: list[dict]) -> dict:
    partes = []
    for a in anexos:
        if a["tipo"] == "imagem":
            partes.append({"type": "image_url",
                           "image_url": {"url": a["conteudo"]}})
        else:
            partes.append({"type": "text",
                           "text": f"[Anexo: {a['nome']}]\n{a['conteudo'][:60000]}"})
    partes.append({"type": "text", "text": texto or "(sem texto)"})
    return {"role": "user", "content": partes}


# busca na web (OpenRouter): US$ 0,007 por consulta, com até 10 resultados
PLUGIN_WEB = {"plugins": [{"id": "web", "max_results": 5}]}


def conversar(cliente: ai_client.OpenRouterClient, historico: list[dict],
              texto: str, anexos: list[dict],
              web: bool = False) -> tuple[str, list[dict]]:
    """Envia a mensagem. Devolve (resposta, fontes consultadas na web)."""
    msgs = [{"role": "system", "content": SYSTEM}]
    if web:
        msgs.append({"role": "system", "content":
                     "Você tem busca na web ativada nesta mensagem. Use os "
                     "resultados para responder com informação atual e cite "
                     "de onde tirou. Se os resultados não responderem, diga "
                     "que não encontrou em vez de supor."})
    for m in historico[-20:]:
        msgs.append({"role": m["role"], "content": m["texto"]})
    msgs.append(_mensagem_usuario(texto, anexos))

    msg = cliente._chat(msgs, extras=PLUGIN_WEB if web else None,
                        completo=True)
    fontes = []
    for a in (msg.get("annotations") or []) if web else []:
        if a.get("type") == "url_citation":
            c = a.get("url_citation", {})
            fontes.append({"titulo": c.get("title") or c.get("url", ""),
                           "url": c.get("url", "")})
    # sem repetir a mesma fonte
    vistos, unicas = set(), []
    for f in fontes:
        if f["url"] and f["url"] not in vistos:
            vistos.add(f["url"])
            unicas.append(f)
    return msg.get("content") or "", unicas


def extrair_acao(resposta: str) -> dict | None:
    m = re.search(r'\{[^{}]*"acao"\s*:.*\}', resposta, flags=re.S)
    if not m:
        return None
    try:
        return json.loads(m.group(0))
    except json.JSONDecodeError:
        return None


# ------------------------------------------------------------- imagens

def gerar_imagem(cliente: ai_client.OpenRouterClient, prompt: str,
                 pasta: str) -> str:
    """Gera uma imagem e devolve o caminho do PNG salvo."""
    r = requests.post(
        ai_client.API_URL, timeout=300,
        headers={"Authorization": f"Bearer {cliente.api_key}",
                 "Content-Type": "application/json",
                 "X-Title": "Formatador de Relatorios"},
        json={"model": MODELO_IMAGEM,
              "messages": [{"role": "user", "content": prompt}],
              "modalities": ["image", "text"]})
    if r.status_code == 402:
        raise ai_client.ApiError("Créditos insuficientes para gerar imagem.")
    r.raise_for_status()
    dados = r.json()
    msg = dados["choices"][0]["message"]
    imagens = msg.get("images") or []
    if not imagens:
        raise ai_client.ApiError("O modelo não devolveu imagem.")
    url = imagens[0]["image_url"]["url"]
    b64 = url.split(",", 1)[1] if url.startswith("data:") else None
    if not b64:
        raise ai_client.ApiError("Formato de imagem inesperado.")
    os.makedirs(pasta, exist_ok=True)
    carimbo = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    destino = os.path.join(pasta, f"imagem_{carimbo}.png")
    with open(destino, "wb") as f:
        f.write(base64.b64decode(b64))
    return destino


# ------------------------------------------------------------ planilhas

def aplicar_operacoes(caminho: str, operacoes: list[dict],
                      destino: str | None = None) -> str:
    """Aplica as operações pedidas e salva uma CÓPIA (nunca sobrescreve)."""
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter, range_boundaries

    wb = load_workbook(caminho)

    def aba(op):
        nome = op.get("planilha")
        return wb[nome] if nome in wb.sheetnames else wb.active

    for op in operacoes:
        tipo = (op.get("tipo") or "").lower()
        ws = aba(op)
        if tipo in ("escrever", "formula"):
            ws[op["celula"]] = op["valor"]
        elif tipo == "preencher_abaixo":
            origem = ws[op["celula"]]
            col, linha = origem.column, origem.row
            formula = str(origem.value or "")
            for l in range(linha + 1, int(op.get("ate_linha", linha)) + 1):
                nova = re.sub(r"([A-Z]+)(\d+)",
                              lambda m: f"{m.group(1)}{int(m.group(2)) + (l - linha)}",
                              formula)
                ws.cell(row=l, column=col).value = nova
        elif tipo == "formato":
            ini_c, ini_l, fim_c, fim_l = range_boundaries(op["intervalo"])
            for l in range(ini_l, fim_l + 1):
                for c in range(ini_c, fim_c + 1):
                    cel = ws.cell(row=l, column=c)
                    if op.get("negrito") or op.get("italico"):
                        cel.font = Font(bold=bool(op.get("negrito")),
                                        italic=bool(op.get("italico")))
                    if op.get("fundo"):
                        cor = str(op["fundo"]).lstrip("#").upper()[:6]
                        cel.fill = PatternFill(start_color=cor,
                                               end_color=cor,
                                               fill_type="solid")
                    if op.get("centralizar"):
                        cel.alignment = Alignment(horizontal="center",
                                                  vertical="center",
                                                  wrap_text=True)
        elif tipo == "numero":
            ini_c, ini_l, fim_c, fim_l = range_boundaries(op["intervalo"])
            for l in range(ini_l, fim_l + 1):
                for c in range(ini_c, fim_c + 1):
                    ws.cell(row=l, column=c).number_format = op["formato"]
        elif tipo == "largura":
            col = op["coluna"]
            col = col if isinstance(col, str) else get_column_letter(int(col))
            ws.column_dimensions[col].width = float(op["valor"])
        elif tipo == "congelar":
            ws.freeze_panes = op.get("celula", "A2")

    if destino is None:
        base, ext = os.path.splitext(caminho)
        destino = f"{base}_EDITADO{ext}"
    wb.save(destino)
    return destino


# --------------------------------------------------- ações de arquivo

def executar_arquivo(acao: dict, anexos: list[dict],
                     pasta_saida: str | None = None) -> tuple[str, str]:
    """Executa a operação pedida. Devolve (mensagem, caminho_gerado)."""
    import arquivos

    tipo = (acao.get("tipo") or "").lower()
    par = acao.get("parametros") or {}
    caminhos = [a["caminho"] for a in anexos if a.get("caminho")]
    if not caminhos:
        return ("Anexe o arquivo e repita o pedido.", "")

    def dos(ext):
        return [c for c in caminhos if c.lower().endswith(ext)]

    pdfs = dos(".pdf")
    imagens = [c for c in caminhos
               if os.path.splitext(c)[1].lower() in IMAGENS]
    docxs, pptxs = dos(".docx"), dos(".pptx")

    def destino(base, sufixo, ext):
        pasta = pasta_saida or os.path.dirname(os.path.abspath(base))
        nome = os.path.splitext(os.path.basename(base))[0]
        return os.path.join(pasta, f"{nome}{sufixo}{ext}")

    if tipo == "juntar_pdf":
        if len(pdfs) < 2:
            return ("Anexe pelo menos dois PDFs para juntar.", "")
        saida = arquivos.pdf_juntar(pdfs, destino(pdfs[0], "_JUNTADO", ".pdf"))
        return (f"PDFs unidos em {os.path.basename(saida)}.", saida)

    if tipo == "dividir_pdf":
        arqs = arquivos.pdf_dividir(pdfs[0], pasta_saida)
        return (f"PDF dividido em {len(arqs)} arquivos, um por página.",
                arqs[0] if arqs else "")

    if tipo == "extrair_paginas":
        saida = arquivos.pdf_extrair_paginas(
            pdfs[0], str(par.get("paginas", "1")),
            destino(pdfs[0], "_PAGINAS", ".pdf"))
        return (f"Páginas {par.get('paginas')} extraídas em "
                f"{os.path.basename(saida)}.", saida)

    if tipo == "girar_pdf":
        saida = arquivos.pdf_girar(
            pdfs[0], int(par.get("graus", 90)), str(par.get("paginas", "")),
            destino(pdfs[0], "_GIRADO", ".pdf"))
        return (f"PDF girado em {os.path.basename(saida)}.", saida)

    if tipo == "proteger_pdf":
        senha = par.get("senha")
        if not senha:
            return ("Qual senha devo usar para proteger o PDF?", "")
        saida = arquivos.pdf_proteger(
            pdfs[0], str(senha), destino(pdfs[0], "_PROTEGIDO", ".pdf"))
        return (f"PDF protegido em {os.path.basename(saida)}. Guarde a "
                f"senha: sem ela o arquivo não abre.", saida)

    if tipo == "pdf_para_docx":
        saida, aviso = arquivos.pdf_para_docx(
            pdfs[0], destino(pdfs[0], "_EDITAVEL", ".docx"))
        return (f"Convertido para {os.path.basename(saida)}.\n\n{aviso}",
                saida)

    if tipo == "pdf_para_imagens":
        arqs = arquivos.pdf_para_imagens(pdfs[0], pasta_saida)
        return (f"{len(arqs)} página(s) salvas como imagem.",
                arqs[0] if arqs else "")

    if tipo == "imagens_para_pdf":
        if not imagens:
            return ("Anexe as imagens que devo juntar no PDF.", "")
        saida = arquivos.imagens_para_pdf(
            imagens, destino(imagens[0], "_DOCUMENTO", ".pdf"))
        return (f"{len(imagens)} imagem(ns) reunidas em "
                f"{os.path.basename(saida)}.", saida)

    if tipo == "docx_para_pdf":
        saida = arquivos.docx_para_pdf(
            docxs[0], destino(docxs[0], "", ".pdf"))
        return (f"PDF gerado: {os.path.basename(saida)}.", saida)

    if tipo == "docx_para_pptx":
        saida = arquivos.docx_para_pptx(
            docxs[0], destino(docxs[0], "", ".pptx"),
            titulo=par.get("titulo"))
        return (f"Apresentação gerada: {os.path.basename(saida)}. "
                f"Confira os slides antes de usar.", saida)

    if tipo == "pptx_para_pdf":
        saida = arquivos.pptx_para_pdf(
            pptxs[0], destino(pptxs[0], "", ".pdf"))
        return (f"PDF gerado: {os.path.basename(saida)}.", saida)

    return (f"Não conheço a operação “{tipo}”.", "")
