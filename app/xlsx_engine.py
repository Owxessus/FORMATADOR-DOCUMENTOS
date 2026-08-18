# -*- coding: utf-8 -*-
"""Motor de planilhas .xlsx: correção ortográfica dos textos.

Saídas:
  NOME_FINAL.xlsx       — planilha com os textos corrigidos
  NOME_ALTERACOES.xlsx  — células alteradas realçadas em amarelo +
                          aba "Alterações" com antes → depois
Números, fórmulas, datas e formatação nunca são tocados (travado em código).
"""
from __future__ import annotations

import os
import re

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

HIGHLIGHT = PatternFill(start_color="FFF3B0", end_color="FFF3B0",
                        fill_type="solid")
MIN_LEN = 4  # células menores que isso não vão à IA (códigos, siglas)


def _is_text_cell(cell) -> bool:
    if cell.data_type != "s":          # somente strings literais
        return False
    v = cell.value
    if not isinstance(v, str) or len(v.strip()) < MIN_LEN:
        return False
    if not re.search(r"[A-Za-zÀ-ú]{3}", v):   # precisa ter palavra de verdade
        return False
    return True


def numbers_preserved(orig: str, corr: str) -> bool:
    return sorted(re.findall(r"\d+", orig)) == sorted(re.findall(r"\d+", corr))


def process(xlsx_path: str, corrector, out_dir: str | None = None,
            extra_instructions: str = "",
            progress=lambda msg: None) -> dict:
    progress("Lendo planilha…")
    wb_final = load_workbook(xlsx_path)      # saída corrigida
    wb_marked = load_workbook(xlsx_path)     # saída com realce

    refs, texts = [], []
    for ws in wb_final.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if _is_text_cell(cell):
                    refs.append((ws.title, cell.row, cell.column))
                    texts.append(cell.value)

    if not texts:
        raise ValueError("Nenhum texto encontrado para revisar na planilha.")

    progress("Corrigindo textos com IA…")
    corrected = corrector(texts, ["cell"] * len(texts), extra_instructions)
    if len(corrected) != len(texts):
        raise ValueError("A correção devolveu número inesperado de células.")

    warnings, changes = [], []
    for (sheet, r, c), orig, corr in zip(refs, texts, corrected):
        corr = corr if corr else orig
        if not numbers_preserved(orig, corr):
            warnings.append(f"Números alterados revertidos em {sheet}!R{r}C{c}")
            corr = orig
        if corr != orig:
            changes.append((sheet, r, c, orig, corr))
            wb_final[sheet].cell(row=r, column=c).value = corr
            mcell = wb_marked[sheet].cell(row=r, column=c)
            mcell.value = corr
            mcell.fill = HIGHLIGHT

    progress("Gerando aba de alterações…")
    log = wb_marked.create_sheet("Alterações")
    header = ["Planilha", "Célula", "Texto original", "Texto corrigido"]
    for j, h in enumerate(header, 1):
        cell = log.cell(row=1, column=j, value=h)
        cell.font = Font(bold=True)
    for i, (sheet, r, c, orig, corr) in enumerate(changes, 2):
        col_letter = wb_marked[sheet].cell(row=r, column=c).coordinate
        log.cell(row=i, column=1, value=sheet)
        log.cell(row=i, column=2, value=col_letter)
        log.cell(row=i, column=3, value=orig)
        log.cell(row=i, column=4, value=corr)
    for col, width in zip("ABCD", (18, 10, 60, 60)):
        log.column_dimensions[col].width = width

    base = os.path.splitext(os.path.basename(xlsx_path))[0]
    out_dir = out_dir or os.path.dirname(os.path.abspath(xlsx_path))
    final_path = os.path.join(out_dir, f"{base}_FINAL.xlsx")
    marked_path = os.path.join(out_dir, f"{base}_ALTERACOES.xlsx")

    progress("Salvando arquivos…")
    wb_final.save(final_path)
    wb_marked.save(marked_path)

    return {"final": final_path, "redline": marked_path,
            "verified": True, "warnings": warnings,
            "paragraphs": len(texts), "changed": len(changes)}
